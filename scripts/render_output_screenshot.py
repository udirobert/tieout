"""Render a sample of the tieout output workbook to a PNG for README/docs."""
from pathlib import Path
import openpyxl
from PIL import Image, ImageDraw, ImageFont

FONT = "/Library/Fonts/Arial Unicode.ttf"
HEADER_BG = "#1a1f2e"
ROW_BG_1 = "#0b0e15"
ROW_BG_2 = "#12161f"
HIGHLIGHT = "#3a1c1c"
TEXT = "#e8ecf7"
MUTED = "#a3b3d6"
LINE = "#2a3142"

SHOW_COLS = [
    "Account Number",
    "Bank reference",
    "Narrative",
    "Pulled Out Sender/Beneficiary",
    "Matched Sender/Beneficiary",
    "Resolved Position",
]


def trunc(s, max_len=60):
    s = str(s) if s is not None else ""
    if len(s) > max_len:
        return s[: max_len - 3] + "..."
    return s


def main():
    out_path = Path("docs/assets/output-screenshot.png")
    wb = openpyxl.load_workbook("/tmp/syndicate-demo/outputs/close-tieout-bank-cp.xlsx")
    ws = wb["Staging Sheet"]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_indices = [headers.index(c) + 1 for c in SHOW_COLS]

    rows = []
    for r in range(2, min(12, ws.max_row + 1)):
        rows.append([ws.cell(r, c).value for c in col_indices])

    # Find rows with empty Matched Sender/Beneficiary (col 5 in SHOW_COLS)
    match_idx = SHOW_COLS.index("Matched Sender/Beneficiary")

    title_font = ImageFont.truetype(FONT, 28)
    header_font = ImageFont.truetype(FONT, 22)
    cell_font = ImageFont.truetype(FONT, 20)

    # Compute column widths
    col_widths = [180, 180, 420, 260, 260, 260]
    for i, h in enumerate(SHOW_COLS):
        bbox = header_font.getbbox(trunc(h))
        w = bbox[2] - bbox[0] + 40
        if w > col_widths[i]:
            col_widths[i] = w
    for row in rows:
        for i, val in enumerate(row):
            bbox = cell_font.getbbox(trunc(val))
            w = bbox[2] - bbox[0] + 40
            if w > col_widths[i]:
                col_widths[i] = w

    row_h = 44
    header_h = 54
    title_h = 60
    width = sum(col_widths) + 2
    height = title_h + header_h + len(rows) * row_h + 2

    img = Image.new("RGB", (width, height), ROW_BG_1)
    draw = ImageDraw.Draw(img)

    # Title
    draw.text((20, 15), "tieout — bank counterparty match output (truncated)", fill=TEXT, font=title_font)

    # Header
    x = 1
    y = title_h
    for i, h in enumerate(SHOW_COLS):
        draw.rectangle([x, y, x + col_widths[i], y + header_h], fill=HEADER_BG, outline=LINE)
        draw.text((x + 12, y + 12), trunc(h, 35), fill=TEXT, font=header_font)
        x += col_widths[i]

    # Rows
    for ri, row in enumerate(rows):
        y = title_h + header_h + ri * row_h
        x = 1
        bg = HIGHLIGHT if not row[match_idx] else (ROW_BG_1 if ri % 2 == 0 else ROW_BG_2)
        for ci, val in enumerate(row):
            draw.rectangle([x, y, x + col_widths[ci], y + row_h], fill=bg, outline=LINE)
            color = MUTED if val is None else TEXT
            draw.text((x + 12, y + 10), trunc(val, 48), fill=color, font=cell_font)
            x += col_widths[ci]

    out_path.parent.mkdir(parents=True, exist_ok=True)
    img.save(out_path)
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
