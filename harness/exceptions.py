"""tieout exception queue — human review for flagged answer cells.

After the agent writes a proposed workbook, this module scans the answer range
for values that need human confirmation (blank, #N/A, Excel errors, review
sentinels like "EXCEPTION", or cells the verifier could not fill). It writes
an `exceptions.json` queue with source-row evidence and provides a CLI review
step that applies only approved writes to the final workbook.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils.cell import coordinate_to_tuple

HARNESS = Path(__file__).resolve().parent
ROOT = HARNESS.parent
sys.path.insert(0, str(ROOT / "research"))
sys.path.insert(0, str(HARNESS))

from parsing import cell_ref, normalize_cell_value  # noqa: E402
from sb import answer_cells  # noqa: E402

_XL_ERR = re.compile(
    r"^#(NAME\?|REF!|VALUE!|DIV/0!|N/A|NULL!|NUM!|GETTING_DATA!|ERR!)", re.I
)

_REVIEW_SENTINELS = {"EXCEPTION", "REVIEW", "FLAG", "PENDING", "HOLD"}


def _is_exception_value(value: object) -> bool:
    if value is None or value == "":
        return True
    if isinstance(value, str):
        if _XL_ERR.match(value):
            return True
        if value.strip().upper() in _REVIEW_SENTINELS:
            return True
    return False


def _row_keys(ws, coord: str) -> list[str]:
    """Collect candidate lookup keys from columns left of the answer cell."""
    row, col = coordinate_to_tuple(coord)
    keys: list[str] = []
    for c in range(1, col):
        v = ws.cell(row=row, column=c).value
        if v is not None:
            s = str(v).strip()
            if s and s not in keys and len(s) > 1:
                keys.append(s)
    if not keys:
        # Fallback: first non-empty value in the row.
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=c).value
            if v is not None:
                s = str(v).strip()
                if s:
                    keys.append(s)
                    break
    return keys[:5]


def _find_evidence(wb, keys: list[str], answer_sheet: str, answer_row: int) -> list[dict]:
    """Search all sheets for rows containing any key; return up to 5 evidence rows."""
    evidence: list[dict] = []
    seen: set[tuple[str, int, str]] = set()
    key_set = set(keys)
    for ws in wb.worksheets:
        for r in range(1, ws.max_row + 1):
            if ws.title == answer_sheet and r == answer_row:
                continue
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                s = str(v).strip()
                if s in key_set:
                    ident = (ws.title, r, s)
                    if ident not in seen:
                        seen.add(ident)
                        evidence.append({"sheet": ws.title, "row": r, "key": s})
                    if len(evidence) >= 5:
                        return evidence
    return evidence


def _reason_for(value: object, ref: str, written: dict[str, object]) -> str:
    if ref not in written:
        return "missing answer cell"
    if value is None or value == "":
        return "empty answer cell"
    if isinstance(value, str) and _XL_ERR.match(value):
        return "excel error"
    if isinstance(value, str) and value.strip().upper() in _REVIEW_SENTINELS:
        return "review sentinel"
    return "pending review"


def build_exceptions(task: dict, out_path: Path, status: str, reason: str, written: dict) -> list[dict]:
    """Return exception entries for answer cells that need human review."""
    out_path = Path(out_path)
    out_wb = openpyxl.load_workbook(out_path, data_only=True)
    init_wb = openpyxl.load_workbook(task["init_xlsx"], data_only=True)

    exceptions: list[dict] = []
    for sheet, coord in answer_cells(task, out_wb):
        ws = out_wb[sheet] if sheet and sheet in out_wb.sheetnames else out_wb.active
        ref = cell_ref(sheet or ws.title, coord)
        value = normalize_cell_value(ws[coord].value)

        if ref in written and not _is_exception_value(value):
            continue

        exc_reason = _reason_for(value, ref, written) if _is_exception_value(value) else reason
        if ref not in written and "missing answer" not in (exc_reason or ""):
            exc_reason = "missing answer cell"

        keys = _row_keys(ws, coord)
        row, _ = coordinate_to_tuple(coord)
        evidence = _find_evidence(init_wb, keys, sheet or ws.title, row)
        if not evidence:
            evidence = [{"sheet": sheet or ws.title, "row": row, "key": keys[0] if keys else coord}]

        exceptions.append(
            {
                "cell": ref,
                "reason": exc_reason,
                "evidence_rows": evidence,
                "proposed_value": "" if value is None else value,
                "status": "pending",
            }
        )

    out_wb.close()
    init_wb.close()
    return exceptions


def write_exceptions(
    out_dir: Path,
    task: dict,
    status: str,
    reason: str,
    info: dict | None,
    out_path: Path,
) -> dict:
    """Write per-task and aggregate exceptions.json after a pipeline run."""
    out_dir = Path(out_dir)
    exceptions_dir = out_dir / "exceptions"
    exceptions_dir.mkdir(parents=True, exist_ok=True)

    written = info.get("written", {}) if info else {}
    exceptions = build_exceptions(task, out_path, status, reason, written)

    payload = {
        "task_id": task["id"],
        "init_xlsx": str(task["init_xlsx"]),
        "output": str(out_path),
        "status": status,
        "reason": reason,
        "exceptions": exceptions,
    }

    task_file = exceptions_dir / f"{task['id']}.json"
    task_file.write_text(json.dumps(payload, indent=2, default=str) + "\n", encoding="utf-8")

    # Aggregate file across all tasks in the run.
    agg_file = out_dir / "exceptions.json"
    agg: list[dict] = []
    if agg_file.exists():
        try:
            data = json.loads(agg_file.read_text(encoding="utf-8"))
            agg = data if isinstance(data, list) else [data]
        except Exception:
            agg = []
    agg = [x for x in agg if x.get("task_id") != task["id"]]
    agg.append(payload)
    agg_file.write_text(json.dumps(agg, indent=2, default=str) + "\n", encoding="utf-8")

    return payload


def _load_exceptions(path: Path) -> list[dict]:
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    return data if isinstance(data, list) else [data]


def _set_cell(wb, ref: str, value: object) -> None:
    if "!" in ref:
        sheet, coord = ref.rsplit("!", 1)
        ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    else:
        ws = wb.active
        coord = ref
    ws[coord] = value


def _apply_decisions(payloads: list[dict], decisions: dict[str, dict[str, str]]) -> None:
    """Apply review decisions: approved exceptions keep proposed value, rejected revert to init."""
    for payload in payloads:
        task_id = payload["task_id"]
        init = Path(payload["init_xlsx"])
        out = Path(payload["output"])
        approved = decisions.get(task_id, {})

        if not out.exists():
            shutil.copy(init, out)
        wb = openpyxl.load_workbook(out, data_only=False)
        init_wb = openpyxl.load_workbook(init, data_only=True)
        for exc in payload["exceptions"]:
            cell = exc["cell"]
            is_approved = exc["status"] == "approved" and approved.get(cell) == "approved"
            if is_approved:
                _set_cell(wb, cell, exc["proposed_value"])
            else:
                # Rejected/pending: revert to the init value.
                if "!" in cell:
                    sheet, coord = cell.rsplit("!", 1)
                    ws = init_wb[sheet] if sheet in init_wb.sheetnames else init_wb.active
                else:
                    ws = init_wb.active
                    coord = cell
                _set_cell(wb, cell, ws[coord].value)
        wb.save(out)
        init_wb.close()


def _prompt(msg: str) -> str:
    return input(msg).strip().lower()


def review_exceptions(path: Path | str, approve_all: bool = False, reject_all: bool = False) -> None:
    """Interactive CLI: review and approve/reject each exception, then rewrite output."""
    path = Path(path)
    payloads = _load_exceptions(path)
    decisions: dict[str, dict[str, str]] = {}

    try:
        for payload in payloads:
            task_id = payload["task_id"]
            print(f"\nTask: {task_id}  ({len(payload['exceptions'])} exception(s))")
            task_decisions: dict[str, str] = {}
            for exc in payload["exceptions"]:
                print(f"  {exc['cell']} | {exc['reason']} | proposed: {exc['proposed_value']!r}")
                for ev in exc["evidence_rows"]:
                    print(f"    evidence: {ev['sheet']} row {ev['row']} (key={ev['key']})")

                if approve_all:
                    ans = "y"
                elif reject_all:
                    ans = "n"
                else:
                    ans = _prompt("  approve? [y/n/s/a(all)/q(quit)]: ")
                if ans == "a":
                    approve_all = True
                    ans = "y"
                if ans == "q":
                    break
                if ans == "y":
                    exc["status"] = "approved"
                    task_decisions[exc["cell"]] = "approved"
                elif ans == "n":
                    exc["status"] = "rejected"
                    task_decisions[exc["cell"]] = "rejected"
                else:
                    exc["status"] = "pending"
            decisions[task_id] = task_decisions
    except (EOFError, KeyboardInterrupt):
        pass

    _apply_decisions(payloads, decisions)

    # Persist reviewed statuses back to the exceptions file.
    agg = payloads if len(payloads) > 1 else payloads[0]
    path.write_text(json.dumps(agg, indent=2, default=str) + "\n", encoding="utf-8")
    print("\nReview complete. Output workbook(s) updated with approved exceptions only.")


def main() -> None:
    raw = sys.argv[1:]
    if raw and raw[0] == "review":
        raw = raw[1:]
    parser = argparse.ArgumentParser(description="Review tieout exception queue")
    parser.add_argument("path", help="path to exceptions.json or a task exception file")
    parser.add_argument("--approve-all", action="store_true", help="approve every exception")
    parser.add_argument("--reject-all", action="store_true", help="reject every exception")
    args = parser.parse_args(raw)
    review_exceptions(args.path, approve_all=args.approve_all, reject_all=args.reject_all)


if __name__ == "__main__":
    main()
