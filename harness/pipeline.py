"""tieout pipeline — values-first + codegen. Entry: container + local runs.

  python harness/pipeline.py --dataset-dir /data --out-dir /out [--ids ...]

Default model: tinker:Qwen/Qwen3.8-27B. Sheet-level goes codegen first (write
openpyxl → sandbox exec → read back → repair); cell-level goes values-first,
then one codegen attempt. Values-first is always the last-resort fallback.
Missing line / missing file = 0; on total failure we copy the init workbook.
"""

import argparse
import asyncio
import datetime
import json
import os
import re
import shutil
import sys
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
HARNESS = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT / "research"))
sys.path.insert(0, str(HARNESS))
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402
from openpyxl.cell.cell import MergedCell  # noqa: E402
from sb import answer_cells, load_dataset  # noqa: E402

from adapters import make_completer  # noqa: E402
from executor import run_snippet  # noqa: E402
from parsing import cell_ref, normalize_cell_value, parse_answer, parse_code  # noqa: E402
from prompts import (  # noqa: E402
    SYSTEM_VALUES,
    build_codegen_prompt,
    build_codegen_repair_prompt,
    build_repair_prompt,
    build_values_prompt,
    classify,
    codegen_system,
)
from serializer import serialize_task_workbook  # noqa: E402
from tracer import append_trace  # noqa: E402
from verifier import MAX_ATTEMPTS, postcheck_soffice, sanity_check  # noqa: E402

_ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}([ T]\d{2}:\d{2}(:\d{2})?)?$")
DEFAULT_MODEL = "tinker:Qwen/Qwen3.8-27B"


def _coerce(value, target_ws, coord):
    """Dates to serials; numeric strings to numbers; strip text. Scorer: 2dp / ''==empty."""
    if value is None:
        return ""
    if isinstance(value, str) and _ISO_DATE.match(value.strip()):
        numfmt = target_ws[coord].number_format or ""
        if any(tok in numfmt for tok in ("yy", "dd", "d/", "mm", "h:")) or "T" in value:
            try:
                return datetime.datetime.fromisoformat(value.strip().replace(" ", "T"))
            except ValueError:
                pass
    return normalize_cell_value(value)


def _set_cell(ws, coord, value) -> None:
    """Write even when coord sits inside a merge (baseline crash: MergedCell read-only)."""
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                ws.cell(rng.min_row, rng.min_col).value = value
                return
        return
    ws[coord] = value


def _lookup_value(cells: dict, sheet, coord, ws_title, graded_pairs):
    """Resolve a model cell against graded (sheet, coord), allowing bare B6 when unique."""
    for key in (cell_ref(sheet, coord), cell_ref(ws_title, coord), coord):
        if key not in cells:
            continue
        if key == coord:
            n = sum(1 for s, c in graded_pairs if c == coord)
            if n > 1:
                continue
        return True, cells[key]
    return False, None


def write_output(task: dict, answer, out_path: Path) -> dict:
    cells = {}
    for c in answer.cells:
        cells[cell_ref(c.sheet, c.cell)] = c.value
        cells.setdefault(c.cell, c.value)
    shutil.copy(task["init_xlsx"], out_path)
    wb = openpyxl.load_workbook(out_path)
    graded_pairs = list(answer_cells(task, wb))
    written = {}
    graded_refs = []
    for sheet, coord in graded_pairs:
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        ref = cell_ref(sheet or ws.title, coord)
        graded_refs.append(ref)
        found, v = _lookup_value(cells, sheet, coord, ws.title, graded_pairs)
        if found:
            v = _coerce("" if v is None else v, ws, coord)
            _set_cell(ws, coord, v)
            written[ref] = v
    wb.save(out_path)
    return {"graded": graded_refs, "written": written}


def read_graded(task: dict, out_path: Path) -> dict:
    """Re-read written cells after codegen; coerce ISO dates in place."""
    wb = openpyxl.load_workbook(out_path)
    written = {}
    graded_refs = []
    dirty = False
    for sheet, coord in answer_cells(task, wb):
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        ref = cell_ref(sheet or ws.title, coord)
        v = _coerce(ws[coord].value, ws, coord)
        if v != ws[coord].value:
            _set_cell(ws, coord, v)
            dirty = True
        written[ref] = v
        graded_refs.append(ref)
    if dirty:
        wb.save(out_path)
    return {"graded": graded_refs, "written": written}


def _ensure_output(task: dict, out: Path) -> None:
    if not out.exists():
        shutil.copy(task["init_xlsx"], out)


def _new_trace(complete, step: int, prompt: str) -> dict:
    return {
        "step": step,
        "model": complete.model_name,
        "prompt": prompt,
        "response": None,
        "input_tokens": None,
        "output_tokens": None,
        "latency_ms": None,
        "error": None,
        "tool": None,
        "tool_input": None,
        "tool_output": None,
    }


def _accept(task: dict, out: Path, graded: list[str], written: dict) -> tuple[bool, str]:
    ok, reason = sanity_check(graded, written)
    if not ok:
        return ok, reason
    ok2, reason2 = postcheck_soffice(task, out)
    if not ok2:
        return False, reason2
    return True, reason


async def run_values_loop(ctx: dict, attempts: int) -> str:
    task, out, complete, out_dir = ctx["task"], ctx["out"], ctx["complete"], ctx["out_dir"]
    prompt = build_values_prompt(task, ctx["wb_serialized"])
    last_reply = ""
    status = "error: no values attempt"
    last_reason = ""
    for attempt in range(attempts):
        ctx["step"] += 1
        trace = _new_trace(complete, ctx["step"], prompt)
        started = time.time()
        try:
            text, in_tok, out_tok = await complete(prompt, SYSTEM_VALUES)
            trace.update(response=text, input_tokens=in_tok, output_tokens=out_tok)
            answer = parse_answer(text)
            info = await asyncio.to_thread(write_output, task, answer, out)
            ok, reason = _accept(task, out, info["graded"], info["written"])
            last_reason = reason
            if ok:
                trace["latency_ms"] = int((time.time() - started) * 1000)
                append_trace(out_dir, task["id"], trace)
                return "ok"
            status = f"partial: {reason}"
            last_reply = text
        except Exception as e:  # noqa: BLE001 — best guess, never blank
            status = f"error: {type(e).__name__}: {e}"[:200]
            trace["error"] = status
            last_reply = trace.get("response") or ""
        trace["latency_ms"] = int((time.time() - started) * 1000)
        append_trace(out_dir, task["id"], trace)
        if attempt + 1 < attempts and last_reply:
            prompt = build_repair_prompt(
                task,
                last_reply,
                last_reason or status,
                ctx["graded_refs"],
                workbook_text=ctx["wb_serialized"],
            )
    return f"{status} (values, {last_reason})"[:200]


async def run_codegen_loop(ctx: dict, attempts: int) -> str:
    task, out, complete, out_dir = ctx["task"], ctx["out"], ctx["complete"], ctx["out_dir"]
    prompt = build_codegen_prompt(task, ctx["wb_serialized"], ctx["graded_refs"])
    last_code = ""
    last_stdout = last_stderr = ""
    status = "error: no codegen attempt"
    last_reason = ""
    for attempt in range(attempts):
        ctx["step"] += 1
        trace = _new_trace(complete, ctx["step"], prompt)
        started = time.time()
        try:
            text, in_tok, out_tok = await complete(prompt, codegen_system(task))
            trace.update(response=text, input_tokens=in_tok, output_tokens=out_tok)
            code = parse_code(text)
            result = await asyncio.to_thread(
                run_snippet, code, task["init_xlsx"], str(out)
            )
            trace["tool"] = "exec"
            trace["tool_input"] = code[:4000]
            trace["tool_output"] = {
                "ok": result["ok"],
                "stdout": result["stdout"],
                "stderr": result["stderr"],
                "error": result["error"],
                "summary": result["summary"],
            }
            last_code = code
            last_stdout = result["stdout"]
            last_stderr = result["stderr"]
            if not result["ok"]:
                last_reason = result["error"] or "exec failed"
                status = f"error: {last_reason}"[:200]
                trace["error"] = status
            else:
                info = await asyncio.to_thread(read_graded, task, out)
                ok, reason = _accept(task, out, info["graded"], info["written"])
                last_reason = reason
                if ok:
                    trace["latency_ms"] = int((time.time() - started) * 1000)
                    append_trace(out_dir, task["id"], trace)
                    return "ok"
                status = f"partial: {reason}"
        except Exception as e:  # noqa: BLE001 — best guess, never blank
            status = f"error: {type(e).__name__}: {e}"[:200]
            trace["error"] = status
            last_code = last_code or (trace.get("response") or "")
        trace["latency_ms"] = int((time.time() - started) * 1000)
        append_trace(out_dir, task["id"], trace)
        if attempt + 1 < attempts and last_code:
            prompt = build_codegen_repair_prompt(
                task,
                last_code,
                last_reason or status,
                last_stdout,
                last_stderr,
                ctx["graded_refs"],
            )
    return f"{status} (codegen, {last_reason})"[:200]


async def predict_task(
    complete, task: dict, out_dir: Path, sem: asyncio.Semaphore, path: str = "auto"
) -> str:
    out = out_dir / "outputs" / f"{task['id']}.xlsx"
    async with sem:
        wb0 = openpyxl.load_workbook(task["init_xlsx"])
        graded_refs = []
        for sheet, coord in answer_cells(task, wb0):
            ws = wb0[sheet] if sheet and sheet in wb0.sheetnames else wb0.active
            graded_refs.append(cell_ref(sheet or ws.title, coord))
        ctx = {
            "complete": complete,
            "task": task,
            "out": out,
            "out_dir": out_dir,
            "step": 0,
            "wb_serialized": serialize_task_workbook(task),
            "graded_refs": graded_refs,
        }
        if path == "values":
            status = await run_values_loop(ctx, MAX_ATTEMPTS)
        elif path == "codegen":
            status = await run_codegen_loop(ctx, MAX_ATTEMPTS)
        else:
            kind = classify(task)
            if kind == "sheet-level":
                status = await run_codegen_loop(ctx, MAX_ATTEMPTS)
                if status != "ok":
                    fallback = await run_values_loop(ctx, 1)
                    status = (
                        fallback
                        if fallback == "ok"
                        else f"{status}; values-fallback: {fallback}"[:200]
                    )
            else:
                status = await run_values_loop(ctx, MAX_ATTEMPTS)
                if status != "ok":
                    fallback = await run_codegen_loop(ctx, 1)
                    status = (
                        fallback
                        if fallback == "ok"
                        else f"{status}; codegen-fallback: {fallback}"[:200]
                    )
        _ensure_output(task, out)
        return status


def _load_env() -> None:
    for env_path in (ROOT / ".env", ROOT / "research" / ".env"):
        for line in env_path.read_text().splitlines() if env_path.exists() else []:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def _already_predicted(pred_path: Path) -> set[str]:
    done: set[str] = set()
    if not pred_path.exists():
        return done
    for line in pred_path.read_text().splitlines():
        if not line.strip():
            continue
        try:
            done.add(json.loads(line)["id"])
        except (json.JSONDecodeError, KeyError):
            continue
    return done


async def main() -> None:
    args = parse_args()
    _load_env()

    out_dir = Path(args.out_dir)
    if args.fresh:
        for name in ("predictions.jsonl", "run.log"):
            p = out_dir / name
            if p.exists():
                p.write_text("", encoding="utf-8")
        for sub in ("outputs", "traces"):
            shutil.rmtree(out_dir / sub, ignore_errors=True)
    for sub in ("outputs", "traces"):
        (out_dir / sub).mkdir(parents=True, exist_ok=True)
    (out_dir / "run.log").touch(exist_ok=True)

    tasks = load_dataset(args.dataset_dir)
    if args.ids:
        keep = {i.strip() for i in args.ids.split(",") if i.strip()}
        tasks = [t for t in tasks if t["id"] in keep]

    skipped = 0
    if args.resume:
        done = _already_predicted(out_dir / "predictions.jsonl")
        if done:
            before = len(tasks)
            tasks = [t for t in tasks if t["id"] not in done]
            skipped = before - len(tasks)

    complete = make_completer(args.model, temperature=args.temperature)

    def log(line: str) -> None:
        print(line, flush=True)
        with (out_dir / "run.log").open("a") as f:
            f.write(line + "\n")

    log(
        f"model {complete.model_name}  temp {args.temperature}  path {args.path}  "
        f"tasks {len(tasks)}" + (f"  resume-skip {skipped}" if skipped else "")
    )
    sem = asyncio.Semaphore(args.concurrency)

    async def one(task):
        t0 = time.time()
        status = await predict_task(complete, task, out_dir, sem, path=args.path)
        elapsed_s = round(time.time() - t0, 1)
        log(f"{task['id']:<8} {elapsed_s:>6}s  {status}")
        line = {
            "id": task["id"],
            "output": f"outputs/{task['id']}.xlsx",
            "status": status,
            "elapsed_s": elapsed_s,
        }
        with (out_dir / "predictions.jsonl").open("a") as f:
            f.write(json.dumps(line) + "\n")

    run_t0 = time.time()
    await asyncio.gather(*(one(t) for t in tasks))
    log(f"total {round(time.time() - run_t0, 1)}s for {len(tasks)} tasks")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--dataset-dir", required=True)
    p.add_argument("--out-dir", required=True)
    p.add_argument("--ids", help="comma-separated task ids (default: all)")
    p.add_argument("--model", default=DEFAULT_MODEL)
    p.add_argument(
        "--temperature",
        type=float,
        default=0.0,
        help="sampling temperature (0 for graded runs; 0.7 for B rejection sampling)",
    )
    p.add_argument(
        "--path",
        choices=("auto", "values", "codegen"),
        default="auto",
        help="auto = sheet codegen / cell values-first + fallback; values|codegen force one path",
    )
    p.add_argument("--concurrency", type=int, default=4)
    p.add_argument(
        "--resume",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="skip task ids already present in predictions.jsonl (default: true)",
    )
    p.add_argument(
        "--fresh",
        action="store_true",
        help="wipe predictions/outputs/traces/run.log in --out-dir before running",
    )
    return p.parse_args()


if __name__ == "__main__":
    asyncio.run(main())
