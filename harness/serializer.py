"""tieout serializer — upstream serialization + fill-aware + pinned answer range.

Formatting-gated tasks (docs/TAXONOMY.md #4) need cell.fill: plain serialization
drops it. When the instruction mentions highlighting/color, append the highlighted
cells per sheet. The 120x30 / 20k preview can hide the answer range — we always
append a pinned excerpt of graded cells so truncation never drops them.
"""

import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "research"))

import openpyxl  # noqa: E402
from sb import answer_cells, serialize_workbook  # noqa: E402

MAX_WORKBOOK_CHARS = 20000
PIN_LIMIT = 200

_FILL_TRIGGER = re.compile(r"yellow|highlight|shad|fill colou?r|colou?r.*fill", re.I)


def _fill_lines(path: str, max_rows=120, max_cols=30) -> str:
    wb = openpyxl.load_workbook(path)
    out = []
    for ws in wb.worksheets:
        hits = []
        for row in ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row, max_rows),
            max_col=min(ws.max_column, max_cols),
        ):
            for cell in row:
                fg = (
                    getattr(cell.fill, "fgColor", None)
                    if cell.fill and cell.fill.patternType
                    else None
                )
                if fg is not None and (
                    fg.rgb not in (None, "00000000")
                    or fg.theme not in (None, 0)
                    or fg.indexed not in (None, 64)
                ):
                    shown = getattr(fg, "rgb", None) or (
                        f"theme{fg.theme}" if fg.theme not in (None, 0) else fg.indexed
                    )
                    hits.append(f"{cell.coordinate}={shown}")
        if hits:
            out.append(
                f"### Sheet: {ws.title} highlighted cells: {', '.join(hits[:200])}"
            )
    return "\n".join(out)


def _answer_range_excerpt(task: dict) -> str:
    """Pin graded coordinates only — never init values (C: models echoed placeholders)."""
    wb = openpyxl.load_workbook(task["init_xlsx"], data_only=True)
    lines = []
    pairs = list(answer_cells(task, wb))
    for sheet, coord in pairs[:PIN_LIMIT]:
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        lines.append(f"{ws.title}!{coord}")
    header = f"### Answer range (pinned addresses only, {len(pairs)} cells"
    if len(pairs) > PIN_LIMIT:
        header += f", showing first {PIN_LIMIT}"
    header += "; init values omitted)"
    return header + "\n" + "\n".join(lines) if lines else header


def serialize_task_workbook(task: dict) -> str:
    text = serialize_workbook(task["init_xlsx"])
    if _FILL_TRIGGER.search(task.get("instruction", "")):
        extra = _fill_lines(task["init_xlsx"])
        if extra:
            text += "\n\n" + extra
    pinned = _answer_range_excerpt(task)
    note = ""
    overhead = len(pinned) + 120
    if len(text) + overhead > MAX_WORKBOOK_CHARS:
        keep = max(MAX_WORKBOOK_CHARS - overhead, 0)
        text = text[:keep]
        note = (
            "\n\n[TRUNCATED — workbook larger than 20k chars; "
            "answer range cells appended below]\n"
        )
    return text + note + "\n\n" + pinned
