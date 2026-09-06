#!/usr/bin/env python3
"""Build space-light Syndicate demo fixtures from Ylookup anonymised datasets.

Reads from ~/Downloads/Ylookup Hackathon Datasets (or YLOOKUP_DATASETS env).
Writes SpreadsheetBench-compatible tasks under demo/close-tieout/ (~200–500 KB total).

Usage:
  python demo/build_fixtures.py
  python demo/build_fixtures.py --dry-run
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "demo" / "close-tieout"
DEFAULT_SRC = Path.home() / "Downloads" / "Ylookup Hackathon Datasets"


def _copy_sheet_style(src_ws, dst_ws, min_row, max_row, min_col, max_col):
    for row in src_ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            dc = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dc._style = copy(cell._style)


def _write_task(
    task_id: str,
    instruction: str,
    instruction_type: str,
    answer_sheet: str,
    answer_position: str,
    init_wb: Workbook,
    golden_wb: Workbook | None,
    tasks: list[dict],
) -> None:
    rel = f"spreadsheet/{task_id}"
    folder = OUT / rel
    folder.mkdir(parents=True, exist_ok=True)
    init_path = folder / f"1_{task_id}_init.xlsx"
    init_wb.save(init_path)
    init_wb.close()
    golden_path = None
    if golden_wb is not None:
        golden_path = folder / f"1_{task_id}_golden.xlsx"
        golden_wb.save(golden_path)
        golden_wb.close()
    prompt_path = folder / "prompt.txt"
    prompt_path.write_text(instruction.strip() + "\n", encoding="utf-8")
    tasks.append(
        {
            "id": task_id,
            "instruction": instruction.strip(),
            "spreadsheet_path": rel,
            "instruction_type": instruction_type,
            "answer_position": answer_position,
            "answer_sheet": answer_sheet,
            "data_position": answer_position,
            "source": "ylookup-anonymised",
        }
    )


def build_le_mapping(src: Path, tasks: list[dict], n_rows: int = 20) -> None:
    """Fund admin: map legal entities to target-system LE IDs (Corvus LE ID column)."""
    ref = src / "02-investor-level-gl-to-loader/output/Tranche 1 - reference and verified loader v4c (anonymised).xlsx"
    wb = openpyxl.load_workbook(ref, data_only=False)
    ws = wb["LE Mapping"]

    init = Workbook()
    init.remove(init.active)
    init_ws = init.create_sheet("LE Mapping")
    gold = Workbook()
    gold.remove(gold.active)
    gold_ws = gold.create_sheet("LE Mapping")

    max_row = min(2 + n_rows, ws.max_row)
    for r in range(1, max_row + 1):
        for c in range(1, 7):
            val = ws.cell(r, c).value
            init_ws.cell(r, c, val)
            gold_ws.cell(r, c, val)
        if r >= 3:
            init_ws.cell(r, 5, None)  # blank Corvus LE ID for agent to fill

    instruction = (
        "You are preparing a fund-administration migration upload. On sheet 'LE Mapping', "
        "fill column E (Corvus LE ID) for each legal entity in column B by looking up the "
        "matching Corvus LE name in column D. Use the same numeric ID as in the reference "
        "mapping — each row's ID must match its entity. Do not change headers or other columns."
    )
    _write_task(
        "close-tieout-le-map",
        instruction,
        "Cell-Level Manipulation",
        "LE Mapping",
        f"E3:E{max_row}",
        init,
        gold,
        tasks,
    )
    wb.close()


def build_movements_rec(src: Path, tasks: list[dict], n_rows: int = 12) -> None:
    """Pre-upload reconciliation: flag non-zero net movements as EXCEPTION."""
    ref = src / "02-investor-level-gl-to-loader/output/Tranche 1 - reference and verified loader v4c (anonymised).xlsx"
    wb = openpyxl.load_workbook(ref, data_only=True)
    ws = wb["Movements Rec"]

    init = Workbook()
    init.remove(init.active)
    init_ws = init.create_sheet("Movements Rec")
    gold = Workbook()
    gold.remove(gold.active)
    gold_ws = gold.create_sheet("Movements Rec")

    headers = ["Legal Entity", "Verado II GL Account", "Sum Debits", "Sum Credits", "Net Movement", "Status"]
    for c, h in enumerate(headers, 1):
        init_ws.cell(1, c, h)
        gold_ws.cell(1, c, h)

    row_out = 2
    for r in range(2, ws.max_row + 1):
        if row_out > 1 + n_rows:
            break
        row = [ws.cell(r, c).value for c in range(1, 6)]
        if not any(row):
            continue
        net = row[4] or 0
        for c, val in enumerate(row, 1):
            init_ws.cell(row_out, c, val)
            gold_ws.cell(row_out, c, val)
        status = "OK" if abs(float(net)) < 0.01 else "EXCEPTION"
        init_ws.cell(row_out, 6, None)
        gold_ws.cell(row_out, 6, status)
        row_out += 1

    last = row_out - 1
    instruction = (
        "Pre-upload reconciliation for month-end close. On sheet 'Movements Rec', fill column F "
        "(Status) for each row: write 'OK' if Net Movement (column E) is zero (within rounding), "
        "otherwise write 'EXCEPTION'. Rows with EXCEPTION require controller review before upload. "
        "Do not modify columns A–E."
    )
    _write_task(
        "close-tieout-movements-rec",
        instruction,
        "Cell-Level Manipulation",
        "Movements Rec",
        f"F2:F{last}",
        init,
        gold,
        tasks,
    )
    wb.close()


def build_bank_counterparty(src: Path, tasks: list[dict], n_rows: int = 15) -> None:
    """Treasury: match pulled counterparty strings to vendor master list."""
    ref = src / "01-bank-statements-to-journal-entries/workbook/Bank statement to journal entries - working file (anonymised).xlsx"
    wb = openpyxl.load_workbook(ref, data_only=False)

    init = Workbook()
    init.remove(init.active)
    init_ws = init.create_sheet("Staging Sheet")
    gold = Workbook()
    gold.remove(gold.active)
    gold_ws = gold.create_sheet("Staging Sheet")
    vendor_ws = wb["Vendor Master List"]

    staging = wb["Staging Sheet"]
    # copy header
    for c in range(1, 26):
        v = staging.cell(1, c).value
        init_ws.cell(1, c, v)
        gold_ws.cell(1, c, v)

    row_out = 2
    for r in range(2, staging.max_row + 1):
        if row_out > 1 + n_rows:
            break
        pulled = staging.cell(r, 10).value  # Pulled Out Sender/Beneficiary
        matched = staging.cell(r, 11).value  # Matched Sender/Beneficiary
        if not pulled:
            continue
        for c in range(1, 26):
            val = staging.cell(r, c).value
            init_ws.cell(row_out, c, val)
            gold_ws.cell(row_out, c, val)
        init_ws.cell(row_out, 11, None)  # blank match column
        gold_ws.cell(row_out, 11, matched)
        row_out += 1

    # include vendor master for lookup context (second sheet)
    vend_init = init.create_sheet("Vendor Master List")
    vend_gold = gold.create_sheet("Vendor Master List")
    for r in range(1, min(vendor_ws.max_row + 1, 80)):
        for c in range(1, 3):
            v = vendor_ws.cell(r, c).value
            vend_init.cell(r, c, v)
            vend_gold.cell(r, c, v)

    last = row_out - 1
    instruction = (
        "Bank statement processing for fund treasury close. On sheet 'Staging Sheet', match each "
        "Pulled Out Sender/Beneficiary (column J) to the clean vendor name in column K using the "
        "'Vendor Master List' sheet. Write the matched full vendor name in column K. If no match "
        "exists, leave column K blank — those rows will route to the exception queue for review."
    )
    _write_task(
        "close-tieout-bank-cp",
        instruction,
        "Cell-Level Manipulation",
        "Staging Sheet",
        f"K2:K{last}",
        init,
        gold,
        tasks,
    )
    wb.close()


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", type=Path, default=Path(os.environ.get("YLOOKUP_DATASETS", DEFAULT_SRC)))
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not args.source.is_dir():
        raise SystemExit(f"Source not found: {args.source}\nSet YLOOKUP_DATASETS or install datasets in Downloads.")

    tasks: list[dict] = []
    if args.dry_run:
        print(f"Would build 3 fixtures from {args.source} -> {OUT}")
        return

    if OUT.exists():
        shutil.rmtree(OUT)
    OUT.mkdir(parents=True)

    build_le_mapping(args.source, tasks)
    build_movements_rec(args.source, tasks)
    build_bank_counterparty(args.source, tasks)

    (OUT / "dataset.json").write_text(json.dumps(tasks, indent=2) + "\n", encoding="utf-8")
    total = sum(f.stat().st_size for f in OUT.rglob("*") if f.is_file())
    print(f"Built {len(tasks)} demo tasks under {OUT} ({total // 1024} KB total)")
    for t in tasks:
        print(f"  - {t['id']}")


if __name__ == "__main__":
    main()
