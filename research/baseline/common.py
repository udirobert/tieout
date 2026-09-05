"""Shared by llm_predict.py and tinker_predict.py: prompt, answer schema, output files, traces."""

import asyncio
import json
import os
import re
import shutil
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from pydantic import BaseModel, Field

from sb import answer_cells, load_dataset, serialize_workbook

ENV_PATH = Path(__file__).resolve().parent.parent / ".env"

SYSTEM_PROMPT = (
    "You are a spreadsheet expert. You get a serialized workbook and a user instruction. "
    "Compute the final values the answer range must contain after the instruction is applied. "
    "Return one entry per cell in the answer range. Use null for cells that must be empty. "
    "Return plain values, not formulas."
)
FORMAT_HINT = (
    '\n\nReply with JSON only, no prose, in this shape: '
    '{"cells": [{"cell": "B6", "value": 42}, {"cell": "B7", "value": null}]}'
)


class CellValue(BaseModel):
    cell: str = Field(description="Cell address like A3 or B6")
    value: str | int | float | bool | None = Field(description="Final value for that cell")


class SpreadsheetAnswer(BaseModel):
    cells: list[CellValue]


def load_env(path: Path = ENV_PATH) -> None:
    if not path.exists():
        return
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def parse_ids(text: str | None) -> set[str] | None:
    return {i.strip() for i in text.split(",") if i.strip()} if text else None


def selected_tasks(dataset_dir: Path, ids: set[str] | None) -> list[dict]:
    tasks = load_dataset(dataset_dir)
    return tasks if ids is None else [t for t in tasks if t["id"] in ids]


def build_prompt(task: dict) -> str:
    return (
        f"## Instruction\n{task['instruction']}\n\n"
        f"## Workbook\n{serialize_workbook(task['init_xlsx'])}\n\n"
        f"## Answer range\nSheet: {task.get('answer_sheet') or 'active sheet'}\nCells: {task['answer_position']}\n"
    )


def parse_answer(text: str) -> SpreadsheetAnswer:
    """First {...} block in the reply. Thinking models wrap it in prose or code fences."""
    text = re.sub(r"<think>.*?</think>", "", text, flags=re.S)
    start, end = text.find("{"), text.rfind("}")
    if start < 0 or end < 0:
        raise ValueError(f"no JSON object in reply: {text[:120]!r}")
    return SpreadsheetAnswer.model_validate(json.loads(text[start:end + 1]))


def write_output(task: dict, answer: SpreadsheetAnswer, out_path: Path) -> None:
    cells = {c.cell.upper(): c.value for c in answer.cells}
    shutil.copy(task["init_xlsx"], out_path)
    wb = openpyxl.load_workbook(out_path)
    for sheet, coord in answer_cells(task, wb):
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        if coord in cells:
            ws[coord] = cells[coord]
    wb.save(out_path)


def prepare_out_dir(out_dir: Path) -> None:
    for sub in ("outputs", "traces"):
        shutil.rmtree(out_dir / sub, ignore_errors=True)
        (out_dir / sub).mkdir(parents=True)
    for name in ("predictions.jsonl", "run.log"):
        (out_dir / name).write_text("", encoding="utf-8")


def append_jsonl(path: Path, record: dict) -> None:
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")


def log(out_dir: Path, line: str) -> None:
    print(line, flush=True)
    with (out_dir / "run.log").open("a", encoding="utf-8") as f:
        f.write(line + "\n")


async def predict_task(complete, model: str, task: dict, out_dir: Path) -> str:
    """One model call, one workbook, one trace line. On any failure the init workbook is the output.

    `complete(prompt)` is an async function returning (text, input_tokens, output_tokens).
    """
    out = out_dir / "outputs" / f"{task['id']}.xlsx"
    trace = {"step": 1, "model": model, "prompt": None, "response": None,
             "input_tokens": None, "output_tokens": None, "latency_ms": None, "error": None}
    started = time.time()
    try:
        trace["prompt"] = build_prompt(task)
        text, trace["input_tokens"], trace["output_tokens"] = await complete(trace["prompt"])
        trace["response"] = text
        write_output(task, parse_answer(text), out)
        status = "ok"
    except Exception as e:
        shutil.copy(task["init_xlsx"], out)
        trace["error"] = f"{type(e).__name__}: {e}"[:500]
        status = f"error: {e}"[:200]
    trace["latency_ms"] = int((time.time() - started) * 1000)
    append_jsonl(out_dir / "traces" / f"{task['id']}.jsonl", trace)
    append_jsonl(out_dir / "predictions.jsonl", {"id": task["id"], "output": f"outputs/{task['id']}.xlsx", "status": status})
    return status


async def run(complete, model: str, tasks: list[dict], out_dir: Path, concurrency: int) -> None:
    prepare_out_dir(out_dir)
    log(out_dir, f"model {model}  tasks {len(tasks)}")
    semaphore = asyncio.Semaphore(concurrency)

    async def run_one(task: dict) -> None:
        async with semaphore:
            status = await predict_task(complete, model, task, out_dir)
        log(out_dir, f"{task['id']:<8} {status}")

    await asyncio.gather(*(run_one(task) for task in tasks))
