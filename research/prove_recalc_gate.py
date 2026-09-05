"""Prove soffice recalc-as-gate + hybrid values fallback actually fire.

This Mac has no LibreOffice (sanity-only). Run on tieout-builder:

  cd ~/tieout && python3 research/prove_recalc_gate.py

No Tinker. Fake completer + synthetic / real-task workbooks only.
"""

from __future__ import annotations

import asyncio
import shutil
import sys
import tempfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "research"))
sys.path.insert(0, str(ROOT / "harness"))
sys.path.insert(0, str(ROOT))

from sb import load_answer_values, load_dataset, recalculate, soffice_path  # noqa: E402

from pipeline import predict_task  # noqa: E402
from verifier import is_formula_error_reason, postcheck_soffice  # noqa: E402


def _task(init_xlsx: Path, cell: str = "B6", sheet: str = "Sheet1") -> dict:
    return {
        "id": "gate-synth",
        "init_xlsx": str(init_xlsx),
        "instruction": "Put the answer in B6.",
        "instruction_type": "Sheet-Level Manipulation",
        "answer_position": cell,
        "answer_sheet": sheet,
    }


def _wb_with(path: Path, value, cell: str = "B6") -> None:
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.active[cell] = value
    wb.active["A1"] = 1
    wb.active["A2"] = 2
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _check(name: str, ok: bool, detail: str) -> None:
    mark = "PASS" if ok else "FAIL"
    print(f"  [{mark}] {name}: {detail}")
    if not ok:
        raise SystemExit(f"gate proof failed: {name}")


def prove_soffice_present() -> None:
    exe = soffice_path()
    _check("soffice on PATH", bool(exe), exe or "missing — this is a sanity-only machine")


def prove_postcheck_formulas(tmpdir: Path) -> None:
    cases = [
        ("div0", "=1/0", True, "#DIV/0!"),
        ("ref", "=#REF!", True, "#REF!"),
        ("name", "=UNKNOWNFUNC(1)", True, "#NAME?"),
        ("clean-sum", "=SUM(A1:A2)", False, 3),
        ("clean-literal", 42, False, 42),
    ]
    for name, value, expect_fail, want_val in cases:
        xlsx = tmpdir / f"{name}.xlsx"
        _wb_with(xlsx, value)
        ok, reason = postcheck_soffice(_task(xlsx), xlsx)
        failed = not ok
        with tempfile.TemporaryDirectory() as td:
            recalc_path = recalculate(str(xlsx), td)
            vals = load_answer_values(recalc_path, _task(xlsx))
        got = next(iter(vals.values()))
        val_ok = (
            isinstance(got, str) and str(want_val).lower() in got.lower()
            if expect_fail
            else got == want_val
        )
        _check(
            f"postcheck {name}",
            failed == expect_fail and (is_formula_error_reason(reason) if expect_fail else ok),
            f"ok={ok} reason={reason!r} recalc_value={got!r}",
        )
        _check(f"recalc value {name}", val_ok, f"got {got!r}, want {want_val!r}")


def prove_real_tasks(tmpdir: Path, n: int = 4) -> None:
    dataset = Path.home() / "tieout/research/data/spreadsheetbench_verified_400"
    if not dataset.exists():
        print("  [SKIP] real tasks: dataset not on this machine")
        return
    tasks = load_dataset(dataset)
    sheet = [t for t in tasks if "sheet" in (t.get("instruction_type") or "").lower()]
    picked = sheet[:n]
    _check("real sheet-level tasks loaded", len(picked) >= n, f"{len(picked)} of {len(sheet)}")
    for task in picked:
        src = Path(task["init_xlsx"])
        dest = tmpdir / f"real-{task['id']}.xlsx"
        shutil.copy(src, dest)
        wb = openpyxl.load_workbook(dest)
        from sb import answer_cells

        coords = list(answer_cells(task, wb))
        from openpyxl.cell.cell import MergedCell

        n_wrote = 0
        for sheet_name, coord in coords[:12]:
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            cell = ws[coord]
            if isinstance(cell, MergedCell):
                continue
            ws[coord] = "=1/0"
            n_wrote += 1
        wb.save(dest)
        if n_wrote == 0:
            print(f"  [SKIP] real task {task['id']}: no writable graded cells")
            continue
        ok, reason = postcheck_soffice(task, dest)
        _check(
            f"real task {task['id']} {task.get('instruction_type')}",
            (not ok) and is_formula_error_reason(reason),
            reason,
        )


class _FakeCompleter:
    model_name = "fake-recalc-gate"

    def __init__(self) -> None:
        self.calls: list[str] = []

    async def __call__(self, prompt: str, system: str = "") -> tuple[str, int, int]:
        sys_l = system or ""
        if "self-contained Python" in sys_l or "openpyxl ONLY" in sys_l:
            self.calls.append("codegen")
            text = (
                "```python\n"
                "import openpyxl\n"
                "wb = openpyxl.load_workbook(INIT_XLSX)\n"
                'wb.active["B6"] = "=1/0"\n'
                "wb.save(OUT_XLSX)\n"
                'print(\'SUMMARY_JSON={"status": "ok", "notes": "div0"}\')\n'
                "```"
            )
            return text, 8, 8
        self.calls.append("values")
        return '{"cells": [{"cell": "B6", "value": 7}]}', 4, 4


def prove_hybrid_fallback(tmpdir: Path) -> None:
    init = tmpdir / "init.xlsx"
    _wb_with(init, None)
    out_dir = tmpdir / "hybrid-out"
    (out_dir / "outputs").mkdir(parents=True)
    (out_dir / "traces").mkdir(parents=True)
    fake = _FakeCompleter()
    status = asyncio.run(
        predict_task(fake, _task(init), out_dir, asyncio.Semaphore(1), path="hybrid")
    )
    out = out_dir / "outputs" / "gate-synth.xlsx"
    wb = openpyxl.load_workbook(out)
    written = wb.active["B6"].value
    _check("hybrid status ok after fallback", status == "ok", status)
    _check("codegen ran first", fake.calls[:1] == ["codegen"], str(fake.calls))
    _check("values fallback ran", "values" in fake.calls, str(fake.calls))
    _check("output is values 7, not formula", written == 7, repr(written))
    tpath = out_dir / "traces" / "gate-synth.jsonl"
    lines = [ln for ln in tpath.read_text().splitlines() if ln.strip()] if tpath.exists() else []
    _check("trace recorded both paths", len(lines) >= 2, f"{len(lines)} trace lines at {tpath}")


def main() -> None:
    print("recalc-gate proof (LibreOffice required)")
    prove_soffice_present()
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        prove_postcheck_formulas(tmp)
        prove_real_tasks(tmp)
        prove_hybrid_fallback(tmp)
    print("ALL CHECKS PASSED — error-scan + values fallback fire, not pass-through")


if __name__ == "__main__":
    main()
