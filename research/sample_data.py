"""Rejection-sampling data generation (STaR, owner B).

Runs the harness prompts at temperature ~0.7, n samples per task, and keeps only
trajectories whose output passes the verifier sanity gate (+ soffice post-check).

  python research/sample_data.py --dataset-dir research/data/spreadsheetbench_verified_400 \
      --out-dir research/data/sft --n 8 [--ids ...] [--limit N] [--max-per-task 2]

Paths per task kind:
  cell-level  -> values-first prompt, clean {"cells":[...]} JSON answer kept as-is.
  sheet-level -> codegen prompt (openpyxl script, sandbox exec); on pass, graded
                 values are read back from the executed workbook and converted to
                 the same parser-friendly JSON format (thinking stripped by parse).

Outputs (out-dir):
  trajectories.jsonl  one record per verified sample (resume-aware)
  sft_train.jsonl     deduped, capped, balanced SFT set (built at end of run)
  sft_manifest.json   counts, token totals, model/temp, git sha
"""

import argparse
import asyncio
import datetime
import json
import os
import subprocess
import sys
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "research"))
sys.path.insert(0, str(ROOT / "harness"))
sys.path.insert(0, str(ROOT))

MAX_COMPLETION_LEN = 8000

import openpyxl  # noqa: E402
from sb import answer_cells, load_dataset  # noqa: E402

from adapters import make_completer  # noqa: E402
from executor import run_snippet  # noqa: E402
from parsing import parse_answer, parse_code  # noqa: E402
from pipeline import _accept, _coerce, _load_env, read_graded, write_output  # noqa: E402
from prompts import (  # noqa: E402
    CODEGEN_SYSTEM,
    SYSTEM_VALUES,
    build_codegen_prompt,
    build_values_prompt,
    classify,
)
from serializer import serialize_task_workbook  # noqa: E402


class SkipTask(Exception):
    """Raised when a task is deterministically unpassable (e.g. size gate)."""


# Tasks known to exceed MAX_COMPLETION_LEN even after diff compression.
KNOWN_TOO_LARGE = {"17-35", "23-24", "24-23"}


def _load_skip(out_dir: Path) -> set[str]:
    skip = set(KNOWN_TOO_LARGE)
    f = out_dir / "skip_ids.txt"
    if f.exists():
        skip |= {l.strip() for l in f.read_text().splitlines() if l.strip()}
    return skip


def _record_skip(out_dir: Path, task_id: str) -> None:
    with (out_dir / "skip_ids.txt").open("a") as f:
        f.write(task_id + "\n")


def _clean_answer(answer) -> str:
    """Re-serialize a parsed Answer into the parser-friendly canonical format."""
    cells = [
        {"cell": c.cell, "value": c.value, **({"sheet": c.sheet} if c.sheet else {})}
        for c in answer.cells
    ]
    return json.dumps({"cells": cells}, ensure_ascii=False, separators=(",", ":"))


def _dedupe_key(answer) -> str:
    return json.dumps(
        sorted((c.sheet or "", c.cell, repr(c.value)) for c in answer.cells),
        ensure_ascii=False,
    )


def _values_equal(v1, v2) -> bool:
    if v1 == v2:
        return True
    if isinstance(v1, (int, float)) and isinstance(v2, (int, float)):
        return round(float(v1), 10) == round(float(v2), 10)
    if (v1 is None or v1 == "") and (v2 is None or v2 == ""):
        return True
    return False


def _values_from_written(written: dict, init_values: dict | None = None) -> str:
    """Convert a workbook read-back into a compact diff JSON (all changed cells)."""
    def enc(v):
        if isinstance(v, (datetime.datetime, datetime.date)):
            return v.isoformat()
        return v

    cells = []
    init_values = init_values or {}
    for ref, v in written.items():
        if _values_equal(v, init_values.get(ref)):
            continue
        sheet, _, coord = ref.rpartition("!")
        c = {"cell": coord, "value": enc(v)}
        if sheet:
            c["sheet"] = sheet
        cells.append(c)
    return json.dumps({"cells": cells}, ensure_ascii=False, separators=(",", ":"))


def _already(path: Path) -> set[str]:
    done: set[str] = set()
    if not path.exists():
        return done
    for line in path.read_text().splitlines():
        if not line.strip():
            continue
        try:
            done.add(json.loads(line)["id"])
        except (json.JSONDecodeError, KeyError):
            continue
    return done


async def sample_task(complete, task: dict, args, sem: asyncio.Semaphore, tmp_dir: Path) -> list[dict]:
    """Return verified trajectory records for one task (up to max-per-task)."""
    kind = classify(task)
    wb0 = openpyxl.load_workbook(task["init_xlsx"])
    init_values = {}
    graded_refs = []
    for sheet, coord in answer_cells(task, wb0):
        ws = wb0[sheet] if sheet and sheet in wb0.sheetnames else wb0.active
        ref = f"{sheet or ws.title}!{coord}"
        graded_refs.append(ref)
        init_values[ref] = _coerce(ws[coord].value, ws, coord)
    wb_serialized = serialize_task_workbook(task)
    out_path = tmp_dir / f"{task['id']}.xlsx"

    if kind == "sheet-level":
        prompt = build_codegen_prompt(task, wb_serialized, graded_refs)
        system = CODEGEN_SYSTEM
    else:
        prompt = build_values_prompt(task, wb_serialized)
        system = SYSTEM_VALUES

    kept: list[dict] = []
    seen: set[str] = set()
    tok = [0, 0]

    async def one_sample(idx: int) -> None:
        t0 = time.time()
        async with sem:
            try:
                text, in_tok, out_tok = await complete(prompt, system)
            except Exception as e:  # noqa: BLE001
                print(f"  {task['id']} sample {idx}: {type(e).__name__}: {e}"[:160], flush=True)
                return
        tok[0] += in_tok or 0
        tok[1] += out_tok or 0
        try:
            if kind == "sheet-level":
                code = parse_code(text)
                result = await asyncio.to_thread(run_snippet, code, task["init_xlsx"], str(out_path))
                if not result["ok"]:
                    return
                info = await asyncio.to_thread(read_graded, task, out_path)
                ok, _ = _accept(task, out_path, info["graded"], info["written"])
                if not ok:
                    return
                answer_json = _values_from_written(info["written"], init_values)
                assert len(answer_json) <= MAX_COMPLETION_LEN, f"completion too long: {len(answer_json)} chars"
                key = json.dumps(sorted((r, repr(v)) for r, v in info["written"].items()), default=str)
            else:
                answer = parse_answer(text)
                info = await asyncio.to_thread(write_output, task, answer, out_path)
                ok, _ = _accept(task, out_path, info["graded"], info["written"])
                if not ok:
                    return
                answer_json = _values_from_written(info["written"], init_values)
                assert len(answer_json) <= MAX_COMPLETION_LEN, f"completion too long: {len(answer_json)} chars"
                key = _dedupe_key(answer)
        except SkipTask:
            raise
        except AssertionError as e:
            # Size gate is deterministic: if one sample is too large, all will be.
            print(f"  {task['id']}: size gate -> skipping task ({e})"[:160], flush=True)
            raise SkipTask(str(e)) from e
        except Exception as e:  # noqa: BLE001
            print(f"  {task['id']} sample {idx}: {type(e).__name__}: {e}"[:160], flush=True)
            return
        if key in seen:
            return
        seen.add(key)
        kept.append({
            "id": task["id"],
            "kind": kind,
            "path": "codegen-readback" if kind == "sheet-level" else "values",
            "system": system,
            "prompt": prompt,
            "completion": answer_json,
            "sample_idx": idx,
            "latency_s": round(time.time() - t0, 1),
            "in_tokens": in_tok,
            "out_tokens": out_tok,
        })

    # Run samples concurrently, but stop as soon as the size gate fires — it is
    # deterministic for a task, so all remaining samples would also exceed 8k.
    pending = {asyncio.create_task(one_sample(i)) for i in range(args.n)}
    too_long_streak = 0
    while pending:
        done, pending = await asyncio.wait(pending, return_when=asyncio.FIRST_COMPLETED)
        for fut in done:
            try:
                await fut
            except asyncio.CancelledError:
                pass
            except SkipTask:
                too_long_streak += 1
                if too_long_streak >= 2:
                    for t in pending:
                        t.cancel()
                    pending.clear()
                    break
            except Exception:
                pass
    if too_long_streak:
        _record_skip(tmp_dir.parent, task["id"])
    print(f"{task['id']:<8} {kind:<11} kept {len(kept)}/{args.n}  tok {tok[0]}/{tok[1]}", flush=True)
    return kept[: args.max_per_task]


def build_sft(traj_path: Path, out_dir: Path, max_per_task: int) -> dict:
    records, seen = [], set()
    for line in traj_path.read_text().splitlines():
        if not line.strip():
            continue
        r = json.loads(line)
        key = (r["id"], r["completion"])
        if key in seen:
            continue
        seen.add(key)
        records.append(r)
    per_task: dict[str, list] = {}
    for r in records:
        per_task.setdefault(r["id"], []).append(r)
    balanced = []
    for tid in sorted(per_task):
        balanced.extend(per_task[tid][:max_per_task])
    with (out_dir / "sft_train.jsonl").open("w") as f:
        for r in balanced:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    cell = [r for r in balanced if r["kind"] == "cell-level"]
    sheet = [r for r in balanced if r["kind"] == "sheet-level"]
    manifest = {
        "built_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "git_sha": subprocess.run(
            ["git", "rev-parse", "HEAD"], capture_output=True, text=True, cwd=ROOT
        ).stdout.strip(),
        "model": os.environ.get("SAMPLING_MODEL", "Qwen/Qwen3.8-27B"),
        "temperature": float(os.environ.get("TINKER_TEMPERATURE", "0.7")),
        "n_per_task": os.environ.get("SAMPLING_N", ""),
        "total_verified": len(records),
        "kept": len(balanced),
        "cell_level": len(cell),
        "sheet_level": len(sheet),
        "tasks_with_trajectory": len(per_task),
        "in_tokens": sum(r.get("in_tokens") or 0 for r in records),
        "out_tokens": sum(r.get("out_tokens") or 0 for r in records),
    }
    (out_dir / "sft_manifest.json").write_text(json.dumps(manifest, indent=2) + "\n")
    return manifest


async def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--dataset-dir", required=True)
    p.add_argument("--out-dir", required=True)
    p.add_argument("--ids", help="comma-separated task ids (default: all)")
    p.add_argument("--limit", type=int, default=0, help="first N tasks only (pilot)")
    p.add_argument("--n", type=int, default=8)
    p.add_argument("--concurrency", type=int, default=6)
    p.add_argument("--max-per-task", type=int, default=2)
    p.add_argument("--model", default="tinker:Qwen/Qwen3.8-27B")
    p.add_argument("--no-build", action="store_true", help="skip SFT build at end")
    args = p.parse_args()

    _load_env()
    os.environ.setdefault("TINKER_TEMPERATURE", "0.7")
    os.environ["SAMPLING_N"] = str(args.n)
    os.environ["SAMPLING_MODEL"] = args.model.split(":", 1)[1].split("|")[0]
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    tmp_dir = out_dir / "tmp_outputs"
    tmp_dir.mkdir(exist_ok=True)

    tasks = load_dataset(args.dataset_dir)
    if args.ids:
        keep = {i.strip() for i in args.ids.split(",") if i.strip()}
        tasks = [t for t in tasks if t["id"] in keep]
    if args.limit:
        tasks = tasks[: args.limit]
    done = _already(out_dir / "trajectories.jsonl")
    skip = _load_skip(out_dir)
    before = len(tasks)
    tasks = [t for t in tasks if t["id"] not in done and t["id"] not in skip]
    import random

    random.Random(42).shuffle(tasks)  # fixed seed: interleave cell/sheet kinds
    print(f"sampling {len(tasks)}/{before} tasks  n={args.n}  temp={os.environ['TINKER_TEMPERATURE']}  skip {before - len(tasks)}", flush=True)

    complete = make_completer(args.model, temperature=0.7)
    sem = asyncio.Semaphore(args.concurrency)
    traj_path = out_dir / "trajectories.jsonl"
    totals = [0, 0, 0]

    for task in tasks:
        kept = await sample_task(complete, task, args, sem, tmp_dir)
        with traj_path.open("a") as f:
            for r in kept:
                f.write(json.dumps(r, ensure_ascii=False) + "\n")
        totals[0] += len(kept)
        totals[1] += sum(r.get("in_tokens") or 0 for r in kept)
        totals[2] += sum(r.get("out_tokens") or 0 for r in kept)
        print(f"progress: {totals[0]} trajectories, in {totals[1]}, out {totals[2]}", flush=True)

    if not args.no_build:
        m = build_sft(traj_path, out_dir, args.max_per_task)
        print("manifest: " + json.dumps(m), flush=True)


if __name__ == "__main__":
    asyncio.run(main())
