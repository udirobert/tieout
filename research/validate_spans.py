"""Lossless Span-Encoding Round-Trip Quality Gate (Role C).

Validates that:
1. Spans in SFT completions (e.g. {"span": "A2:A50", "value": X}) expand deterministically.
2. Writing expanded spans through the openpyxl write path and reading them back is 100% lossless.
3. No format drift occurs between B's training representation and A's container write path.

Usage:
  python3 research/validate_spans.py --sft-file research/data/sft/trajectories.jsonl
"""

import argparse
import json
import re
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "research"))
sys.path.insert(0, str(ROOT))

import openpyxl
from openpyxl.utils.cell import range_boundaries, get_column_letter

def expand_span(span_str: str) -> list[str]:
    """Expands bounding box strings like 'A2:B4' or 'Sheet1!A2:A10' into coordinate list."""
    sheet = ""
    coord_range = span_str
    if "!" in span_str:
        sheet, coord_range = span_str.split("!", 1)
    
    coord_range = coord_range.replace("$", "")
    if ":" not in coord_range:
        return [f"{sheet}!{coord_range}" if sheet else coord_range]
    
    min_col, min_row, max_col, max_row = range_boundaries(coord_range)
    cells = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            col_letter = get_column_letter(c)
            coord = f"{col_letter}{r}"
            cells.append(f"{sheet}!{coord}" if sheet else coord)
    return cells

def parse_and_expand_completion(completion_json: dict) -> dict[str, object]:
    """Expands both single cells and span entries into a flat {cell_coord: value} mapping."""
    flat_cells = {}
    
    # Handle {"cells": [...]}
    if "cells" in completion_json:
        for item in completion_json["cells"]:
            if isinstance(item, dict):
                if "span" in item:
                    for c in expand_span(item["span"]):
                        flat_cells[c.upper()] = item.get("value")
                elif "cell" in item:
                    flat_cells[item["cell"].upper().replace("$", "")] = item.get("value")
                    
    # Handle {"spans": [...]}
    if "spans" in completion_json:
        for item in completion_json["spans"]:
            if isinstance(item, dict) and "span" in item:
                for c in expand_span(item["span"]):
                    flat_cells[c.upper()] = item.get("value")
                    
    return flat_cells

from sb import values_equal

def test_round_trip_fidelity(task: dict, raw_completion: str, work_dir: Path) -> tuple[bool, str]:
    """Writes expanded completion to init_xlsx and verifies exact read-back value fidelity."""
    try:
        data = json.loads(raw_completion) if isinstance(raw_completion, str) else raw_completion
    except Exception as e:
        return False, f"JSON parse error: {e}"
    
    expanded = parse_and_expand_completion(data)
    if not expanded:
        return False, "No cells or spans found in completion (empty completion)"
    
    test_out = work_dir / f"test_{task.get('id', 'temp')}.xlsx"
    wb = openpyxl.load_workbook(task["init_xlsx"]) if "init_xlsx" in task and Path(task["init_xlsx"]).exists() else openpyxl.Workbook()
    
    # Write expanded cells
    for coord_key, val in expanded.items():
        sheet_name = None
        coord = coord_key
        if "!" in coord_key:
            sheet_name, coord = coord_key.split("!", 1)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        ws[coord] = val
        
    wb.save(test_out)
    
    # Read back from saved workbook
    wb_read = openpyxl.load_workbook(test_out, data_only=False)
    mismatches = []
    for coord_key, expected_val in expanded.items():
        sheet_name = None
        coord = coord_key
        if "!" in coord_key:
            sheet_name, coord = coord_key.split("!", 1)
        ws = wb_read[sheet_name] if sheet_name and sheet_name in wb_read.sheetnames else wb_read.active
        read_val = ws[coord].value
        
        # Compare using official evaluation values_equal
        if not values_equal(expected_val, read_val):
            mismatches.append(f"{coord_key}: expected {expected_val!r}, got {read_val!r}")
            
    if mismatches:
        return False, f"Read-back mismatch on {len(mismatches)} cells: {mismatches[:3]}"
    
    return True, f"100% Lossless ({len(expanded)} cells expanded and verified)"

def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--sft-file", default="research/data/sft/trajectories.jsonl")
    return p.parse_args()

def main():
    args = parse_args()
    sft_path = Path(args.sft_file)
    if not sft_path.exists():
        print(f"File {sft_path} not found.")
        return 1
    
    records = [json.loads(l) for l in sft_path.read_text().splitlines() if l.strip()]
    print(f"=== LOSSLESS SPAN-ENCODING QUALITY GATE (Role C) ===")
    print(f"Auditing {len(records)} records from {sft_path}...")
    
    passed = 0
    failed = 0
    spans_detected = 0
    
    with tempfile.TemporaryDirectory() as work:
        work_dir = Path(work)
        for idx, r in enumerate(records):
            comp = r.get("completion") or (r.get("messages")[-1]["content"] if "messages" in r else "")
            if "span" in comp:
                spans_detected += 1
            ok, reason = test_round_trip_fidelity(r.get("task", {}), comp, work_dir)
            if ok:
                passed += 1
            else:
                failed += 1
                if failed <= 5:
                    print(f"  [Record {idx}] FAILED: {reason}")
                    
    print(f"\nGate Summary: {passed}/{len(records)} PASSED (Failed: {failed}) | Spans detected: {spans_detected}")
    if failed == 0:
        print("VERDICT: 100% LOSSLESS ROUND-TRIP CONFIRMED.")
        return 0
    else:
        print("VERDICT: GATE REJECTED. Fix span expansion before SFT training.")
        return 1

if __name__ == "__main__":
    sys.exit(main())
